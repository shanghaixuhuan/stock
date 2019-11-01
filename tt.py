import xlwings as xw
import time
import pymongo
from openpyxl import load_workbook

s_col = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X",
         "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ",
         "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ", "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI",
         "BJ", "BK", "BL", "BM", "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ"]

func2 = ["EM_S_SEST_NETPROFITFY1", "EM_S_SEST_NETPROFITFY2", "EM_S_SEST_NETPROFITFY3", "EM_S_SEST_NETPROFITF12",
         "EM_S_SEST_NETPROFITYOY", "EM_S_SEST_NETPROFITGR2", "EM_S_SEST_AVGEPSFY1", "EM_S_SEST_AVGEPSFY2",
         "EM_S_SEST_AVGEPSFY3", "EM_S_SEST_AVGEPSF12", "EM_S_SEST_AVGROEFY1", "EM_S_SEST_AVGROEFY2",
         "EM_S_SEST_AVGROEFY3", "EM_S_SEST_AVGROEF12", "EM_S_SEST_AVGROEYOY", "EM_S_SEST_AVGSALESFY1",
         "EM_S_SEST_AVGSALESFY2","EM_S_SEST_AVGSALESFY3", "EM_S_SEST_AVGSALESF12", "EM_S_SEST_AVGSALESYOY",
         "EM_S_SEST_AVGSALESGR2", "EM_S_SEST_AVGCPSFY1", "EM_S_SEST_AVGCPSFY2", "EM_S_SEST_AVGCPSFY3",
         "EM_S_SEST_AVGCPSF12", "EM_S_SEST_AVGDPSFY1", "EM_S_SEST_AVGDPSFY2", "EM_S_SEST_AVGDPSFY3",
         "EM_S_SEST_AVGDPSF12", "EM_S_SEST_AVGBPSFY1", "EM_S_SEST_AVGBPSFY2", "EM_S_SEST_AVGBPSFY3",
         "EM_S_SEST_AVGBPSF12", "EM_S_SEST_AVGEBITFY1", "EM_S_SEST_AVGEBITFY2", "EM_S_SEST_AVGEBITFY3",
         "EM_S_SEST_AVGEBITF12", "EM_S_SEST_AVGEBITDAFY1", "EM_S_SEST_AVGEBITDAFY2", "EM_S_SEST_AVGEBITDAFY3",
         "EM_S_SEST_AVGEBITDAF12", "EM_S_SEST_AVGOPERATINGPROFITFY1", "EM_S_SEST_AVGOPERATINGPROFITFY2",
         "EM_S_SEST_AVGOPERATINGPROFITFY3", "EM_S_SEST_AVGOPERATINGPROFITF12", "EM_S_SEST_AVGOPERATINGPROFITYOY"]
func3 = ["EM_S_SEST_NETPROFITRATE1W", "EM_S_SEST_NETPROFITRATE4W", "EM_S_SEST_NETPROFITRATE13W",
         "EM_S_SEST_NETPROFITRATE26W"]
date = ["2019-6-1", "2019-7-1", "2019-8-1", "2019-9-1", "2019-10-1"]
year = ["2019", "2020", "2021"]
path = "data.xlsx"

client = "mongodb://localhost:27017/"
db = "shares"
col = "shares"


class Stock:
    def __init__(self, var_path, var_func2, var_func3, var_year):
        self.path = var_path
        self.func2 = var_func2
        self.func3 = var_func3
        self.year = var_year
        self.day = None
        self.func2_num = len(var_func2)
        self.func3_num = len(var_func3)

        self.columns = 0
        self.rows = 0

        self.app = None
        self.wb = None
        self.ws = None

    @staticmethod
    def refreshing(var_dd):
        flag = False
        for i in range(len(var_dd)):
            if "Refreshing" in var_dd[i]:
                flag = True
                break
        return flag

    def load_data(self, var_day):
        self.day = var_day

        self.app = xw.App(visible=True, add_book=False)
        self.wb = self.app.books.add()
        self.ws = self.wb.sheets[0]

        self.ws.range('a1').value = "=EM_SECTOR(\"2000032254\", \"" + var_day + "\")"
        while self.ws["B2"].value is None:
            time.sleep(0.5)
        time.sleep(0.5)
        self.rows = self.ws.api.UsedRange.Rows.count

        loc = 0

        for i in func2:
            self.ws[s_col[loc] + '2:' + s_col[loc] + str(self.rows)].value = "=" + i + "(B2,\"" + var_day + "\")"
            loc += 1

        for i in func3:
            for j in year:
                self.ws[s_col[loc] + '2:' + s_col[loc] + str(self.rows)].value = "=" + i + "(B2,\"" + var_day + "\",\""\
                                                                                 + j + "\" )"
                loc += 1

    def wait_refresh(self):
        self.columns = self.ws.api.UsedRange.Columns.count
        dd = self.ws.range('C2:' + s_col[self.columns - 1] + str(self.rows)).value
        while self.refreshing(dd):
            dd = self.ws.range('C2:' + s_col[self.columns - 1] + str(self.rows)).value

    def save_wb(self):
        self.wb.save(self.path)
        self.wb.close()
        self.app.quit()

    def save_to_mongodb(self, var_client, var_db, var_col):
        my_client = pymongo.MongoClient(var_client)
        my_db = my_client[var_db]
        my_col = my_db[var_col]
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        data_amount = ws.max_row - 1
        for j in range(2, 2+data_amount):
            data = {"stock_name": ws['A' + str(j)].value, "date": self.day, "stock_code": ws['B' + str(j)].value}
            for k in range(self.func2_num):
                data.update({func2[k]: ws[s_col[k] + str(j)].value})
            for k in range(self.func3_num):
                ob = {}
                for l in range(len(year)):
                    ob.update({year[l]: ws[s_col[self.func2_num + k * len(year) + l] + str(j)].value})
                data.update({func3[k]: ob})
            my_col.insert(data)


if __name__ == "__main__":
    s = Stock(path, func2, func3, year)
    for day in date:
        s.load_data(day)
        s.wait_refresh()
        s.save_wb()
        s.save_to_mongodb(client, db, col)

