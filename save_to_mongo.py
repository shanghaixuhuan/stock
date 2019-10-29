import pymongo
from openpyxl import load_workbook

path = "data.xlsx"

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["shares"]
mycol = mydb["shares"]
wb = load_workbook(filename=path,data_only=True)
ws = wb.active
data_amount = ws.max_row - 1
for i in range(2, 2 + data_amount):
    mycol.insert({"name" : ws['A' + str(i)].value, "num" : ws['B' + str(i)].value, "FY1" : ws['C' + str(i)].value,
                  "FY2": ws['D' + str(i)].value, "FY3" : ws['E' + str(i)].value, "F12" : ws['F' + str(i)].value,
                  "YOY" : ws['G' + str(i)].value,})
