from openpyxl import load_workbook
import time

date = time.strftime("%Y-%m-%d", time.localtime())

path = "data.xlsx"
wb = load_workbook(filename=path,data_only=True)
ws = wb.active
data_amount = ws.max_row - 1

for i in range(2, data_amount + 2):
    share_num = ws["B" + str(i)].value
    ws['C' + str(i)] = "=EM_S_SEST_NETPROFITFY1(\"" + share_num + "\",\"" + date + "\")"
    ws['D' + str(i)] = "=EM_S_SEST_NETPROFITFY2(\"" + share_num + "\",\"" + date + "\")"
    ws['E' + str(i)] = "=EM_S_SEST_NETPROFITFY3(\"" + share_num + "\",\"" + date + "\")"
    ws['F' + str(i)] = "=EM_S_SEST_NETPROFITF12(\"" + share_num + "\",\"" + date + "\")"
    ws['G' + str(i)] = "=EM_S_SEST_NETPROFITYOY(\"" + share_num + "\",\"" + date + "\")"

wb.save(filename=path)