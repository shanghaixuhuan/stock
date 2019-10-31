import xlwings as xw
import time
import pymongo
from openpyxl import load_workbook

my_client = pymongo.MongoClient("mongodb://localhost:27017/")
my_db = my_client["shares"]
my_col = my_db["shares"]

path = "data.xlsx"
col = ["C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z",
        "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR",
        "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ","BA","BB","BC","BD","BE","BF","BG","BH","BI","BJ","BK","BL",
        "BM","BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ"]
date = ["2019-6-1"] #,"2019-7-1","2019-8-1","2019-9-1","2019-10-1"
func2 = ["EM_S_SEST_NETPROFITFY1","EM_S_SEST_NETPROFITFY2","EM_S_SEST_NETPROFITFY3",
        "EM_S_SEST_NETPROFITF12","EM_S_SEST_NETPROFITYOY","EM_S_SEST_NETPROFITGR2",
        "EM_S_SEST_AVGEPSFY1","EM_S_SEST_AVGEPSFY2",
        "EM_S_SEST_AVGEPSFY3","EM_S_SEST_AVGEPSF12","EM_S_SEST_AVGROEFY1",
        "EM_S_SEST_AVGROEFY2","EM_S_SEST_AVGROEFY3","EM_S_SEST_AVGROEF12",
        "EM_S_SEST_AVGROEYOY","EM_S_SEST_AVGSALESFY1","EM_S_SEST_AVGSALESFY2",
        "EM_S_SEST_AVGSALESFY3","EM_S_SEST_AVGSALESF12","EM_S_SEST_AVGSALESYOY",
        "EM_S_SEST_AVGSALESGR2","EM_S_SEST_AVGCPSFY1","EM_S_SEST_AVGCPSFY2",
        "EM_S_SEST_AVGCPSFY3","EM_S_SEST_AVGCPSF12","EM_S_SEST_AVGDPSFY1",
        "EM_S_SEST_AVGDPSFY2","EM_S_SEST_AVGDPSFY3","EM_S_SEST_AVGDPSF12",
        "EM_S_SEST_AVGBPSFY1","EM_S_SEST_AVGBPSFY2","EM_S_SEST_AVGBPSFY3",
        "EM_S_SEST_AVGBPSF12","EM_S_SEST_AVGEBITFY1","EM_S_SEST_AVGEBITFY2",
        "EM_S_SEST_AVGEBITFY3","EM_S_SEST_AVGEBITF12","EM_S_SEST_AVGEBITDAFY1",
        "EM_S_SEST_AVGEBITDAFY2","EM_S_SEST_AVGEBITDAFY3","EM_S_SEST_AVGEBITDAF12",
        "EM_S_SEST_AVGOPERATINGPROFITFY1","EM_S_SEST_AVGOPERATINGPROFITFY2",
        "EM_S_SEST_AVGOPERATINGPROFITFY3","EM_S_SEST_AVGOPERATINGPROFITF12",
        "EM_S_SEST_AVGOPERATINGPROFITYOY"]
func3 = ["EM_S_SEST_NETPROFITRATE1W","EM_S_SEST_NETPROFITRATE4W","EM_S_SEST_NETPROFITRATE13W","EM_S_SEST_NETPROFITRATE26W"]
year = ["2019","2020","2021"]

func2_num = len(func2)
func3_num = len(func3)

#"EM_S_SEST_NETPROFITRATE1W","EM_S_SEST_NETPROFITRATE4W","EM_S_SEST_NETPROFITRATE13W","EM_S_SEST_NETPROFITRATE26W",

# for i in range(2, 2 + data_num):
#     if((i-1) % 500 == 0):
#         print("第" + str(i-1) + "条数据完成导入")
#     share_num = ws["B" + str(i)].value
#     ws['C' + str(i)].value = "=EM_S_SEST_NETPROFITFY1(\"" + share_num + "\",\"" + date + "\")"
#     ws['D' + str(i)].value = "=EM_S_SEST_NETPROFITFY2(\"" + share_num + "\",\"" + date + "\")"
#     ws['E' + str(i)].value = "=EM_S_SEST_NETPROFITFY3(\"" + share_num + "\",\"" + date + "\")"
#     ws['F' + str(i)].value = "=EM_S_SEST_NETPROFITF12(\"" + share_num + "\",\"" + date + "\")"
#     ws['G' + str(i)].value = "=EM_S_SEST_NETPROFITYOY(\"" + share_num + "\",\"" + date + "\")"

for i in date:
    app = xw.App(visible=True, add_book=False)
    wb = app.books.add()
    ws = wb.sheets[0]

    start = time.time()
    ws.range('a1').value = "=EM_SECTOR(\"2000032254\", \"" + i + "\")"
    time.sleep(10)
    rows = ws["A1048576"].end('up').row

    loc = 0

    for j in func2:
        ws[col[loc] + '2:' + col[loc] + str(rows)].value = "=" + j + "(B2,\"" + i + "\")"
        loc += 1

    for j in func3:
        for k in year:
            ws[col[loc] + '2:' + col[loc] + str(rows)].value = "=" + j + "(B2,\"" + i + "\",\"" + k + "\" )"
            loc += 1

    loc = 0
    while True:
        if loc == func2_num + len(year) * func3_num:
            break
        elif ws[col[loc] + str(rows)].value == "Refreshing":
            time.sleep(1)
        else:
            loc += 1

    wb.save(path)
    wb.close()
    app.quit()

    end = time.time()
    print(end-start)

    wb = load_workbook(filename=path,data_only=True)
    ws = wb.active
    data_amount = ws.max_row - 1
    for j in range(2, 2+data_amount):
        data = {"stock_name": ws['A' + str(j)].value, "date": i, "stock_code": ws['B' + str(j)].value}
        for k in range(func2_num):
            data.update({func2[k]: ws[col[k] + str(j)].value})
        for k in range(func3_num):
            ob = {}
            for l in range(len(year)):
                ob.update({year[l]: ws[col[func2_num + k * len(year) + l] + str(j)].value})
            data.update({func3[k]: ob})
        my_col.insert(data)


# while ws['C' + str(rows)].value == "Refreshing" or ws['D' + str(rows)].value == "Refreshing" or ws['E' + str(rows)].value == "Refreshing" or \
#         ws['F' + str(rows)].value == "Refreshing" or ws['G' + str(rows)].value == "Refreshing":
#     time.sleep(1)
#
# wb.save(path)
# wb.close()
# app.quit()
#
# print("相关数据完成导入")
#

# wb = load_workbook(filename=path,data_only=True)
# ws = wb.active
# data_amount = ws.max_row - 1
# for i in range(2, 2 + data_amount):
#     mycol.insert({"stock_name" : ws['A' + str(i)].value, "date" : date, "stock_code" : ws['B' + str(i)].value,
#                   "FY1" : ws['C' + str(i)].value, "FY2": ws['D' + str(i)].value, "FY3" : ws['E' + str(i)].value,
#                   "F12" : ws['F' + str(i)].value, "YOY" : ws['G' + str(i)].value,})
#
# print("导入MongoDB完成")